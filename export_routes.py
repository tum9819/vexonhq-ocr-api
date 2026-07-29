"""
VEXONHQ Phase 9 — Export Routes
================================
Endpoints:
    GET /export/category-summary?month=YYYY-MM   → Excel: รายรับ/รายจ่ายต่อ category
    GET /export/daybook?month=YYYY-MM            → Excel: รายการธุรกรรมทั้งหมด
    GET /export/pnd3?month=YYYY-MM               → Excel: ภ.ง.ด. 3 (musician fees + freelancers)
    GET /export/zip-bundle?month=YYYY-MM         → ZIP: รวม 3 ไฟล์ข้างต้น
"""

from __future__ import annotations

import calendar
import io
import logging
import os
import re
import zipfile
from datetime import date, datetime
from typing import Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from auth_routes import _require_admin_role
from bkk import bkk_now, bkk_today
from export_readiness import build_readiness, canonical_fingerprint

try:
    from main import get_db_conn  # type: ignore
except ImportError:
    import psycopg2
    def get_db_conn():
        return psycopg2.connect(os.environ["DATABASE_URL"])

logger = logging.getLogger("export")
router = APIRouter(prefix="/export", tags=["export"])


# ─── Style constants ───────────────────────────────────────────────────────────

EMERALD   = "1D6F42"   # dark green header
EMERALD_L = "E8F5E9"   # light green fill
AMBER_L   = "FFF8E1"
RED_L     = "FFEBEE"
GRAY_H    = "F5F5F5"   # alternate row

FONT_HEADER = Font(name="TH Sarabun New", bold=True, color="FFFFFF", size=12)
FONT_TITLE  = Font(name="TH Sarabun New", bold=True, size=14)
FONT_BODY   = Font(name="TH Sarabun New", size=11)
FONT_BOLD   = Font(name="TH Sarabun New", bold=True, size=11)
FONT_SMALL  = Font(name="TH Sarabun New", size=10, color="888888")

THIN = Side(style="thin", color="CCCCCC")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HEADER = PatternFill("solid", fgColor=EMERALD)
FILL_GRAY   = PatternFill("solid", fgColor=GRAY_H)
FILL_GREEN  = PatternFill("solid", fgColor=EMERALD_L)
FILL_AMBER  = PatternFill("solid", fgColor=AMBER_L)
FILL_RED    = PatternFill("solid", fgColor=RED_L)

CENTER  = Alignment(horizontal="center", vertical="center")
RIGHT   = Alignment(horizontal="right",  vertical="center")
LEFT    = Alignment(horizontal="left",   vertical="center")
WRAP    = Alignment(wrap_text=True, vertical="center")


# ─── Helpers ───────────────────────────────────────────────────────────────────

def _rows_to_dicts(cur) -> list[dict]:
    if not cur.description:
        return []
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def _month_range(month: str) -> tuple[date, date]:
    """'YYYY-MM' → (first_day, last_day)"""
    try:
        y, m = int(month[:4]), int(month[5:7])
    except Exception:
        raise HTTPException(400, f"month must be YYYY-MM, got: {month!r}")
    last = calendar.monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, last)


def _readiness_month_range(month: str) -> tuple[date, date]:
    """Strict YYYY-MM parser for the readiness contract only."""
    if re.fullmatch(r"[0-9]{4}-(0[1-9]|1[0-2])", month) is None:
        raise HTTPException(400, f"month must be YYYY-MM, got: {month!r}")
    y, m = int(month[:4]), int(month[5:7])
    try:
        first = date(y, m, 1)
    except ValueError:
        raise HTTPException(400, f"month must be YYYY-MM, got: {month!r}")
    return first, date(y, m, calendar.monthrange(y, m)[1])


def _month_label_th(month: str) -> str:
    """'2026-05' → 'พฤษภาคม 2569'"""
    MONTHS_TH = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
    ]
    y, m = int(month[:4]), int(month[5:7])
    return f"{MONTHS_TH[m]} {y + 543}"


def _fmt_num(n) -> str:
    if n is None:
        return "-"
    try:
        return f"{float(n):,.2f}"
    except Exception:
        return str(n)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, cols: list[str], row: int = 1):
    for c, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[row].height = 22


def _data_cell(ws, row: int, col: int, value, align=LEFT, bold=False, fill=None, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FONT_BOLD if bold else FONT_BODY
    cell.alignment = align
    cell.border = BORDER_THIN
    if fill:
        cell.fill = fill
    if num_format:
        cell.number_format = num_format
    return cell


def _excel_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Sheet builders ────────────────────────────────────────────────────────────

def _build_category_summary(month: str) -> openpyxl.Workbook:
    """สรุปรายรับ/รายจ่ายต่อ category พร้อม budget comparison"""
    first, last = _month_range(month)
    label = _month_label_th(month)

    wb = openpyxl.Workbook()

    # ── Sheet 1: รายจ่าย ──────────────────────────────────────────────────────
    ws_exp = wb.active
    ws_exp.title = "รายจ่ายต่อหมวด"

    # Title
    ws_exp.merge_cells("A1:G1")
    t = ws_exp.cell(row=1, column=1, value=f"สรุปรายจ่ายต่อหมวดหมู่ — {label}")
    t.font = FONT_TITLE
    t.alignment = CENTER
    ws_exp.row_dimensions[1].height = 28

    # Sub-title
    ws_exp.merge_cells("A2:G2")
    s = ws_exp.cell(row=2, column=1, value=f"ช่วงวันที่: {first.strftime('%d/%m/%Y')} – {last.strftime('%d/%m/%Y')}   |   ออกรายงาน: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.font = FONT_SMALL
    s.alignment = CENTER

    # Header row 3
    _header_row(ws_exp, ["#", "หมวดหมู่", "ชื่อ (EN)", "งบตั้งไว้ (฿)", "Actual (฿)", "ผลต่าง (฿)", "% ใช้"], row=3)

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            # Budget status
            cur.execute(
                """SELECT b.category_code,
                          COALESCE(ec.name_th, b.category_code) AS name_th,
                          COALESCE(ec.name_en, '') AS name_en,
                          b.amount AS budget_amount,
                          COALESCE(a.actual_amount, 0) AS actual_amount,
                          COALESCE(a.actual_amount, 0) - b.amount AS variance,
                          CASE WHEN b.amount = 0 THEN NULL
                               ELSE ROUND((COALESCE(a.actual_amount, 0) / b.amount) * 100, 1)
                          END AS pct_used
                   FROM public.budget_targets b
                   LEFT JOIN public.expense_categories ec ON ec.code = b.category_code
                   LEFT JOIN (
                       SELECT TO_CHAR(entry_date, 'YYYY-MM') AS month,
                              category_code,
                              SUM(amount) AS actual_amount
                       FROM public.v_daybook_pnl
                       WHERE direction = 'expense' AND category_code IS NOT NULL
                       GROUP BY 1, 2
                   ) a ON a.month = b.month AND a.category_code = b.category_code
                   WHERE b.month = %s
                   ORDER BY actual_amount DESC NULLS LAST""",
                (month,),
            )
            budget_rows = _rows_to_dicts(cur)

            # Categories with actual spend (no budget)
            cur.execute(
                """SELECT d.category_code,
                          COALESCE(ec.name_th, d.category_code) AS name_th,
                          COALESCE(ec.name_en, '') AS name_en,
                          SUM(d.amount) AS actual_amount
                   FROM public.v_daybook_pnl d
                   LEFT JOIN public.expense_categories ec ON ec.code = d.category_code
                   WHERE d.direction = 'expense'
                     AND d.category_code IS NOT NULL
                     AND TO_CHAR(d.entry_date, 'YYYY-MM') = %s
                     AND d.category_code NOT IN (
                         SELECT category_code FROM public.budget_targets WHERE month = %s
                     )
                   GROUP BY 1, 2, 3
                   ORDER BY actual_amount DESC""",
                (month, month),
            )
            no_budget_rows = _rows_to_dicts(cur)

            # FA-016: uncategorized spend must appear so the sheet total ties to the
            # daybook expense total (auditors cross-foot the two files).
            cur.execute(
                """SELECT COUNT(*) AS cnt, COALESCE(SUM(amount), 0) AS total
                   FROM public.v_daybook_pnl
                   WHERE direction = 'expense'
                     AND category_code IS NULL
                     AND TO_CHAR(entry_date, 'YYYY-MM') = %s""",
                (month,),
            )
            _u = cur.fetchone()
            uncat_count, uncat_total = int(_u[0]), float(_u[1])
    finally:
        conn.close()

    row = 4
    total_budget = 0.0
    total_actual = 0.0

    for i, r in enumerate(budget_rows, 1):
        pct = float(r["pct_used"]) if r["pct_used"] is not None else None
        budget = float(r["budget_amount"])
        actual = float(r["actual_amount"])
        variance = actual - budget
        total_budget += budget
        total_actual += actual

        fill = None
        if pct is not None:
            if pct >= 100:
                fill = FILL_RED
            elif pct >= 90:
                fill = FILL_AMBER

        _data_cell(ws_exp, row, 1, i, align=CENTER, fill=fill)
        _data_cell(ws_exp, row, 2, r["name_th"], fill=fill)
        _data_cell(ws_exp, row, 3, r["name_en"], fill=fill)
        _data_cell(ws_exp, row, 4, budget, align=RIGHT, num_format='#,##0.00', fill=fill)
        _data_cell(ws_exp, row, 5, actual, align=RIGHT, num_format='#,##0.00', fill=fill)
        _data_cell(ws_exp, row, 6, variance, align=RIGHT, num_format='#,##0.00', fill=fill)
        _data_cell(ws_exp, row, 7, pct, align=RIGHT, num_format='0.0"%"', fill=fill)
        ws_exp.row_dimensions[row].height = 18
        row += 1

    # No-budget rows
    if no_budget_rows:
        ws_exp.merge_cells(f"A{row}:G{row}")
        lbl = ws_exp.cell(row=row, column=1, value="— รายจ่ายที่ยังไม่ได้ตั้งงบ —")
        lbl.font = FONT_SMALL
        lbl.alignment = CENTER
        row += 1
        for r in no_budget_rows:
            actual = float(r["actual_amount"])
            total_actual += actual
            _data_cell(ws_exp, row, 1, "", align=CENTER, fill=FILL_GRAY)
            _data_cell(ws_exp, row, 2, r["name_th"], fill=FILL_GRAY)
            _data_cell(ws_exp, row, 3, r["name_en"], fill=FILL_GRAY)
            _data_cell(ws_exp, row, 4, None, align=CENTER, fill=FILL_GRAY)
            _data_cell(ws_exp, row, 5, actual, align=RIGHT, num_format='#,##0.00', fill=FILL_GRAY)
            _data_cell(ws_exp, row, 6, None, align=CENTER, fill=FILL_GRAY)
            _data_cell(ws_exp, row, 7, None, align=CENTER, fill=FILL_GRAY)
            ws_exp.row_dimensions[row].height = 18
            row += 1

    # FA-016: uncategorized bucket row (before total) so the total ties to daybook.
    if uncat_total > 0:
        _data_cell(ws_exp, row, 1, "", align=CENTER, fill=FILL_AMBER)
        _data_cell(ws_exp, row, 2, f"ไม่ระบุหมวด ({uncat_count} รายการ)", fill=FILL_AMBER)
        _data_cell(ws_exp, row, 3, "Uncategorized", fill=FILL_AMBER)
        _data_cell(ws_exp, row, 4, None, align=CENTER, fill=FILL_AMBER)
        _data_cell(ws_exp, row, 5, uncat_total, align=RIGHT, num_format='#,##0.00', fill=FILL_AMBER)
        _data_cell(ws_exp, row, 6, None, align=CENTER, fill=FILL_AMBER)
        _data_cell(ws_exp, row, 7, None, align=CENTER, fill=FILL_AMBER)
        ws_exp.row_dimensions[row].height = 18
        total_actual += uncat_total
        row += 1

    # Total row
    t1 = ws_exp.cell(row=row, column=1, value="รวมทั้งหมด")
    t1.font = FONT_BOLD
    t1.fill = FILL_GREEN
    t1.alignment = CENTER
    ws_exp.merge_cells(f"A{row}:C{row}")
    _data_cell(ws_exp, row, 4, total_budget, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws_exp, row, 5, total_actual, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws_exp, row, 6, total_actual - total_budget, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws_exp, row, 7, None, fill=FILL_GREEN)
    ws_exp.row_dimensions[row].height = 22

    _set_col_widths(ws_exp, [5, 28, 22, 16, 16, 16, 10])

    # ── Sheet 2: รายรับ ───────────────────────────────────────────────────────
    ws_inc = wb.create_sheet("รายรับต่อหมวด")
    ws_inc.merge_cells("A1:E1")
    t2 = ws_inc.cell(row=1, column=1, value=f"สรุปรายรับต่อหมวดหมู่ — {label}")
    t2.font = FONT_TITLE
    t2.alignment = CENTER
    ws_inc.row_dimensions[1].height = 28

    _header_row(ws_inc, ["#", "หมวดหมู่", "ชื่อ (EN)", "จำนวนครั้ง", "รวม (฿)"], row=2)

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT COALESCE(d.category_code, d.source) AS category_code,
                          COALESCE(ec.name_th, d.category_code,
                            CASE d.source
                              WHEN 'pos_sale'    THEN 'ยอดขาย POS'
                              WHEN 'ar_payment'  THEN 'รับชำระ AR'
                              WHEN 'manual'      THEN 'บันทึกรายรับ'
                              ELSE d.source
                            END
                          ) AS name_th,
                          COALESCE(ec.name_en, '') AS name_en,
                          COUNT(*) AS cnt,
                          SUM(d.amount) AS total_amount
                   FROM public.v_daybook_pnl d
                   LEFT JOIN public.expense_categories ec ON ec.code = d.category_code
                   WHERE d.direction = 'income'
                     AND TO_CHAR(d.entry_date, 'YYYY-MM') = %s
                   GROUP BY 1, 2, 3
                   ORDER BY total_amount DESC""",
                (month,),
            )
            inc_rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    row2 = 3
    total_inc = 0.0
    for i, r in enumerate(inc_rows, 1):
        fill = FILL_GRAY if i % 2 == 0 else None
        total_inc += float(r["total_amount"])
        _data_cell(ws_inc, row2, 1, i, align=CENTER, fill=fill)
        _data_cell(ws_inc, row2, 2, r["name_th"], fill=fill)
        _data_cell(ws_inc, row2, 3, r["name_en"], fill=fill)
        _data_cell(ws_inc, row2, 4, int(r["cnt"]), align=CENTER, fill=fill)
        _data_cell(ws_inc, row2, 5, float(r["total_amount"]), align=RIGHT, num_format='#,##0.00', fill=fill)
        ws_inc.row_dimensions[row2].height = 18
        row2 += 1

    # Total
    ti = ws_inc.cell(row=row2, column=1, value="รวมรายรับ")
    ti.font = FONT_BOLD
    ti.fill = FILL_GREEN
    ti.alignment = CENTER
    ws_inc.merge_cells(f"A{row2}:D{row2}")
    _data_cell(ws_inc, row2, 5, total_inc, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    ws_inc.row_dimensions[row2].height = 22

    _set_col_widths(ws_inc, [5, 28, 22, 12, 16])

    return wb


def _build_daybook(month: str) -> openpyxl.Workbook:
    """รายการธุรกรรมทั้งหมดในเดือน (สมุดรายวัน)"""
    first, last = _month_range(month)
    label = _month_label_th(month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "สมุดรายวัน"

    # Title
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value=f"สมุดรายวัน — {label}")
    t.font = FONT_TITLE
    t.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    s = ws.cell(row=2, column=1, value=f"ช่วงวันที่: {first.strftime('%d/%m/%Y')} – {last.strftime('%d/%m/%Y')}   |   ออกรายงาน: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    s.font = FONT_SMALL
    s.alignment = CENTER

    _header_row(ws, ["วันที่", "ประเภท", "หมวดหมู่", "รายละเอียด", "แหล่งข้อมูล", "รายรับ (฿)", "รายจ่าย (฿)", "สาขา"], row=3)

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT d.entry_date,
                          d.direction,
                          COALESCE(ec.name_th, d.category_code, 'ไม่ระบุ') AS category_name,
                          COALESCE(d.label, d.counterparty, '') AS detail,
                          d.source,
                          CASE WHEN d.direction = 'income'  THEN d.amount ELSE NULL END AS income,
                          CASE WHEN d.direction = 'expense' THEN d.amount ELSE NULL END AS expense,
                          d.branch_code
                   FROM public.v_daybook_pnl d
                   LEFT JOIN public.expense_categories ec ON ec.code = d.category_code
                   WHERE d.entry_date BETWEEN %s AND %s
                   ORDER BY d.entry_date, d.direction DESC, d.amount DESC""",
                (first, last),
            )
            rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    row = 4
    total_inc = 0.0
    total_exp = 0.0
    prev_date = None

    for i, r in enumerate(rows):
        d = r["entry_date"]
        date_str = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)

        # Group separator
        if d != prev_date and prev_date is not None:
            sep_fill = PatternFill("solid", fgColor="EEEEEE")
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = sep_fill
            ws.row_dimensions[row].height = 4
            row += 1

        fill = None
        if r["direction"] == "income":
            fill = FILL_GREEN
            if r["income"]:
                total_inc += float(r["income"])
        else:
            if r["expense"]:
                total_exp += float(r["expense"])

        _data_cell(ws, row, 1, date_str, align=CENTER, fill=fill)
        _data_cell(ws, row, 2, "รายรับ" if r["direction"] == "income" else "รายจ่าย", align=CENTER, fill=fill)
        _data_cell(ws, row, 3, r["category_name"], fill=fill)
        _data_cell(ws, row, 4, r["detail"], fill=fill)
        _data_cell(ws, row, 5, r["source"], align=CENTER, fill=fill)
        _data_cell(ws, row, 6, float(r["income"]) if r["income"] else None, align=RIGHT, num_format='#,##0.00', fill=fill)
        _data_cell(ws, row, 7, float(r["expense"]) if r["expense"] else None, align=RIGHT, num_format='#,##0.00', fill=fill)
        _data_cell(ws, row, 8, r["branch_code"] or "", align=CENTER, fill=fill)
        ws.row_dimensions[row].height = 17
        row += 1
        prev_date = d

    # Summary row
    ws.merge_cells(f"A{row}:E{row}")
    sm = ws.cell(row=row, column=1, value="รวมทั้งเดือน")
    sm.font = FONT_BOLD
    sm.fill = FILL_GREEN
    sm.alignment = CENTER
    _data_cell(ws, row, 6, total_inc, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws, row, 7, total_exp, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws, row, 8, None, fill=FILL_GREEN)
    ws.row_dimensions[row].height = 22

    # Net row
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    nm = ws.cell(row=row, column=1, value="กำไร/ขาดทุนสุทธิ")
    nm.font = FONT_BOLD
    nm.fill = FILL_GREEN
    nm.alignment = CENTER
    net_fill = FILL_GREEN if total_inc >= total_exp else FILL_RED
    _data_cell(ws, row, 6, total_inc - total_exp, align=RIGHT, num_format='#,##0.00', bold=True, fill=net_fill)
    ws.merge_cells(f"G{row}:H{row}")
    ws.cell(row=row, column=7).fill = net_fill
    ws.row_dimensions[row].height = 22

    _set_col_widths(ws, [12, 10, 22, 36, 14, 16, 16, 18])

    # Freeze header rows
    ws.freeze_panes = "A4"

    return wb


def _build_pnd3(month: str) -> openpyxl.Workbook:
    """ภ.ง.ด. 3 — รายชื่อผู้รับเงินที่ต้องหักภาษี ณ ที่จ่าย (บุคคลธรรมดา)
    สำหรับค่าดนตรี / freelancer ที่ต้องหัก 3%"""
    first, last = _month_range(month)
    label = _month_label_th(month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ภ.ง.ด.3"

    # Title block
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="แบบแสดงรายการภาษีเงินได้หัก ณ ที่จ่าย (ภ.ง.ด. 3)")
    t.font = Font(name="TH Sarabun New", bold=True, size=15)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"เดือน: {label}   |   ผู้จ่ายเงิน: ร้านสถานีหม่าล่า  เลขที่ 255/4 ถ.พุทธมณฑลสาย 2 แขวงศาลาธรรมสพน์ เขตทวีวัฒนา กรุงเทพมหานคร 10170")
    ws.cell(row=2, column=1).font = FONT_BODY
    ws.cell(row=2, column=1).alignment = CENTER
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:H3")
    ws.cell(row=3, column=1, value=f"ช่วงวันที่: {first.strftime('%d/%m/%Y')} – {last.strftime('%d/%m/%Y')}")
    ws.cell(row=3, column=1).font = FONT_SMALL
    ws.cell(row=3, column=1).alignment = CENTER

    _header_row(ws, [
        "ลำดับ", "วันที่จ่าย", "ชื่อผู้รับ", "เลขประจำตัวผู้เสียภาษี",
        "ประเภทเงินได้", "อัตราภาษี", "ยอดเงิน (฿)", "ภาษีที่หัก (฿)"
    ], row=4)

    # Single source of truth for which categories are ภ.ง.ด.3 + their per-category
    # WHT rate (audit #1: this generator + yearly + /tax/wht-summary must agree).
    from tax_routes import WHT_RULES  # noqa: PLC0415
    wht_cats = list(WHT_RULES.keys())

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            # ดึงรายการที่ต้องหัก ณ ที่จ่าย (musician_fee/rent/service_fee) จาก v_daybook_pnl
            cur.execute(
                """SELECT entry_date,
                          COALESCE(label, counterparty, 'ไม่ระบุชื่อ') AS name,
                          amount,
                          category_code,
                          source
                   FROM public.v_daybook_pnl
                   WHERE direction = 'expense'
                     AND entry_date BETWEEN %s AND %s
                     AND category_code = ANY(%s)
                   ORDER BY entry_date, amount""",
                (first, last, wht_cats),
            )
            pnd_rows = _rows_to_dicts(cur)

            # PNL-4: best-effort prefill of payee tax-id by EXACT (normalized) name
            # match against counterparties. Exact-match ONLY — a wrong tax-id on a
            # สรรพากร filing is worse than a blank, so never fuzzy-match.
            cur.execute(
                """SELECT lower(btrim(name)) AS k, tax_id
                   FROM public.counterparties
                   WHERE tax_id IS NOT NULL AND btrim(tax_id) <> '' AND is_active"""
            )
            cp_taxid = {k: t for k, t in cur.fetchall() if k}
    finally:
        conn.close()

    row = 5
    total_amount = 0.0
    total_tax = 0.0

    for i, r in enumerate(pnd_rows, 1):
        fill = FILL_GRAY if i % 2 == 0 else None
        d = r["entry_date"]
        date_str = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)
        amount = float(r["amount"])
        rule = WHT_RULES.get(r["category_code"],
                             {"label": "ค่าบริการ", "section": "มาตรา 40(8)", "wht_pct": 3.0})
        pct = float(rule["wht_pct"])
        tax = round(amount * pct / 100, 2)   # per-category rate (rent=5%, others=3%) — not a flat 3%
        total_amount += amount
        total_tax += tax

        _data_cell(ws, row, 1, i, align=CENTER, fill=fill)
        _data_cell(ws, row, 2, date_str, align=CENTER, fill=fill)
        _data_cell(ws, row, 3, r["name"], fill=fill)
        tid = cp_taxid.get((r["name"] or "").strip().lower(), "")  # PNL-4: exact name match, else blank
        _data_cell(ws, row, 4, tid, align=CENTER, fill=fill)
        _data_cell(ws, row, 5, f'{rule["label"]} - {rule["section"]}', fill=fill)
        _data_cell(ws, row, 6, f"{pct:g}%", align=CENTER, fill=fill)
        _data_cell(ws, row, 7, amount, align=RIGHT, num_format='#,##0.00', fill=fill)
        _data_cell(ws, row, 8, tax, align=RIGHT, num_format='#,##0.00', fill=fill)
        ws.row_dimensions[row].height = 18
        row += 1

    if len(pnd_rows) == 0:
        ws.merge_cells(f"A{row}:H{row}")
        empty = ws.cell(row=row, column=1, value="ไม่มีรายการภาษีหัก ณ ที่จ่ายในเดือนนี้")
        empty.font = FONT_SMALL
        empty.alignment = CENTER
        row += 1

    # Total
    ws.merge_cells(f"A{row}:F{row}")
    tl = ws.cell(row=row, column=1, value="รวม")
    tl.font = FONT_BOLD
    tl.fill = FILL_GREEN
    tl.alignment = CENTER
    _data_cell(ws, row, 7, total_amount, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws, row, 8, total_tax, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    ws.row_dimensions[row].height = 22

    # Note
    row += 2
    note = ws.cell(row=row, column=1, value="หมายเหตุ: เลขประจำตัวผู้เสียภาษีที่เติมให้อัตโนมัติมาจากการจับคู่ชื่อแบบตรงตัวเท่านั้น — โปรดตรวจสอบทุกแถวและกรอกช่องที่ว่างก่อนยื่นสรรพากร")
    note.font = Font(name="TH Sarabun New", size=10, italic=True, color="CC0000")
    ws.merge_cells(f"A{row}:H{row}")

    _set_col_widths(ws, [6, 12, 28, 18, 24, 10, 16, 14])

    return wb


def _build_commission_breakdown(month: str) -> openpyxl.Workbook:
    """Commission Breakdown — ยอดขาย Grab/Lineman + Commission หัก"""
    first, last = _month_range(month)
    label = _month_label_th(month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commission Breakdown"

    # Title
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="รายงานการหักค่าคอมมิชชัน")
    t.font = Font(name="TH Sarabun New", bold=True, size=15)
    t.alignment = CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value=f"เดือน: {label}   |   ร้านสถานีหม่าล่า")
    ws.cell(row=2, column=1).font = FONT_BODY
    ws.cell(row=2, column=1).alignment = CENTER

    # Riders Summary Table
    _header_row(ws, [
        "Platform", "ยอดขาย Gross", "ค่าคอมมิชชัน", "ส่วนลดร้าน", "อัตรา %", "Net Payout", "Orders"
    ], row=4)

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT
                     platform,
                     SUM(gross_sales)::numeric AS gross,
                     SUM(net_payout)::numeric AS net,
                     SUM(ABS(gp_amount))::numeric AS commission,
                     SUM(ABS(promo_store))::numeric AS promo,
                     SUM(order_count)::int AS total_orders,
                     COUNT(*) AS delivery_days
                   FROM public.rider_deliveries
                   WHERE delivery_date BETWEEN %s AND %s
                   GROUP BY platform
                   ORDER BY platform""",
                (first, last),
            )
            rider_rows = _rows_to_dicts(cur)
    finally:
        conn.close()

    row = 5
    grand_gross = 0.0
    grand_commission = 0.0
    grand_promo = 0.0
    grand_net = 0.0

    for r in rider_rows:
        gross = float(r["gross"] or 0)
        commission = float(r["commission"] or 0)
        promo = float(r["promo"] or 0)
        net = float(r["net"] or 0)
        orders = int(r["total_orders"] or 0)
        rate = (commission / gross * 100) if gross > 0 else 0.0

        platform_label = "Grab" if r["platform"] == "grab" else "Lineman"

        _data_cell(ws, row, 1, platform_label, bold=True)
        _data_cell(ws, row, 2, gross, align=RIGHT, num_format='#,##0.00')
        _data_cell(ws, row, 3, commission, align=RIGHT, num_format='#,##0.00', fill=FILL_AMBER)
        _data_cell(ws, row, 4, promo, align=RIGHT, num_format='#,##0.00')
        _data_cell(ws, row, 5, f"{rate:.2f}%", align=CENTER)
        _data_cell(ws, row, 6, net, align=RIGHT, num_format='#,##0.00')
        _data_cell(ws, row, 7, orders, align=CENTER)

        grand_gross += gross
        grand_commission += commission
        grand_promo += promo
        grand_net += net
        row += 1

    # Totals
    row += 1
    ws.merge_cells(f"A{row}:A{row}")
    tl = ws.cell(row=row, column=1, value="รวม")
    tl.font = FONT_BOLD
    tl.fill = FILL_GREEN
    tl.alignment = CENTER

    _data_cell(ws, row, 2, grand_gross, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws, row, 3, grand_commission, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)
    _data_cell(ws, row, 4, grand_promo, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)

    grand_rate = (grand_commission / grand_gross * 100) if grand_gross > 0 else 0.0
    _data_cell(ws, row, 5, f"{grand_rate:.2f}%", align=CENTER, bold=True, fill=FILL_GREEN)
    _data_cell(ws, row, 6, grand_net, align=RIGHT, num_format='#,##0.00', bold=True, fill=FILL_GREEN)

    ws.row_dimensions[row].height = 22

    _set_col_widths(ws, [16, 18, 18, 18, 12, 18, 12])

    return wb


# ─── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/category-summary")
def export_category_summary(month: str = Query(..., description="YYYY-MM")):
    """ดาวน์โหลด Excel สรุปรายรับ/รายจ่ายต่อ category พร้อม budget"""
    wb = _build_category_summary(month)
    data = _excel_bytes(wb)
    fname = f"category_summary_{month}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/daybook")
def export_daybook(month: str = Query(..., description="YYYY-MM")):
    """ดาวน์โหลด Excel สมุดรายวันทั้งเดือน"""
    wb = _build_daybook(month)
    data = _excel_bytes(wb)
    fname = f"daybook_{month}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/pnd3")
def export_pnd3(month: str = Query(..., description="YYYY-MM")):
    """ดาวน์โหลด Excel ภ.ง.ด. 3 (ค่าดนตรี + freelancer หัก 3%)"""
    wb = _build_pnd3(month)
    data = _excel_bytes(wb)
    fname = f"pnd3_{month}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/commission-breakdown")
def export_commission_breakdown(month: str = Query(..., description="YYYY-MM")):
    """ดาวน์โหลด Excel รายงานการหักค่าคอมมิชชัน Grab/Lineman"""
    wb = _build_commission_breakdown(month)
    data = _excel_bytes(wb)
    fname = f"commission_breakdown_{month}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/zip-bundle")
def export_zip_bundle(month: str = Query(..., description="YYYY-MM")):
    """ดาวน์โหลด ZIP รวม 3 ไฟล์: category-summary, daybook, pnd3"""
    label = _month_label_th(month)

    category_wb = _build_category_summary(month)
    daybook_wb  = _build_daybook(month)
    pnd3_wb     = _build_pnd3(month)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"category_summary_{month}.xlsx", _excel_bytes(category_wb))
        zf.writestr(f"daybook_{month}.xlsx",          _excel_bytes(daybook_wb))
        zf.writestr(f"pnd3_{month}.xlsx",             _excel_bytes(pnd3_wb))

    zip_buf.seek(0)
    fname = f"VEXONHQ_export_{month}.zip"
    return Response(
        content=zip_buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/summary")
def export_summary(month: str = Query(..., description="YYYY-MM")):
    """สรุปจำนวนรายการแต่ละ export สำหรับแสดงหน้า UI ก่อนดาวน์โหลด"""
    first, last = _month_range(month)

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            # daybook stats
            cur.execute(
                """SELECT COUNT(*) AS cnt,
                          COALESCE(SUM(CASE WHEN direction='income'  THEN amount ELSE 0 END),0) AS total_income,
                          COALESCE(SUM(CASE WHEN direction='expense' THEN amount ELSE 0 END),0) AS total_expense
                   FROM public.v_daybook_pnl
                   WHERE entry_date BETWEEN %s AND %s""",
                (first, last),
            )
            r = cur.fetchone()
            daybook_rows, total_income, total_expense = int(r[0]), float(r[1]), float(r[2])

            # pnd3 stats — use the SAME per-category WHT_RULES as the other 4 generators.
            # (audit AUD-TAX-01: this preview used flat 3% + a phantom-category/amount
            # heuristic AGENTS #19 had removed, so it under-reported rent's 5% WHT and
            # disagreed with the actual pnd3 export.)
            from tax_routes import WHT_RULES  # local import avoids any router import-order issue
            _wht_keys = list(WHT_RULES.keys())
            _wht_case = " ".join("WHEN category_code = %s THEN amount * %s" for _ in _wht_keys)
            _wht_params: list = []
            for _k in _wht_keys:
                _wht_params.extend([_k, WHT_RULES[_k]["wht_pct"] / 100.0])
            cur.execute(
                f"""SELECT COUNT(*) AS cnt,
                          COALESCE(SUM(CASE {_wht_case} ELSE 0 END), 0) AS total_wht
                   FROM public.v_daybook_pnl
                   WHERE direction = 'expense'
                     AND entry_date BETWEEN %s AND %s
                     AND category_code = ANY(%s)""",
                tuple(_wht_params) + (first, last, _wht_keys),
            )
            r2 = cur.fetchone()
            pnd3_rows, total_wht = int(r2[0]), float(r2[1])

            # category summary stats
            cur.execute(
                """SELECT COUNT(DISTINCT category_code) AS cats,
                          COALESCE(SUM(amount), 0) AS total_spend
                   FROM public.v_daybook_pnl
                   WHERE direction = 'expense'
                     AND category_code IS NOT NULL
                     AND entry_date BETWEEN %s AND %s""",
                (first, last),
            )
            r3 = cur.fetchone()
            cat_count, total_spend = int(r3[0]), float(r3[1])

            # FA-016: surface uncategorized spend so the UI can warn before export.
            cur.execute(
                """SELECT COUNT(*), COALESCE(SUM(amount), 0)
                   FROM public.v_daybook_pnl
                   WHERE direction = 'expense'
                     AND category_code IS NULL
                     AND entry_date BETWEEN %s AND %s""",
                (first, last),
            )
            r3u = cur.fetchone()
            uncat_count, uncat_total = int(r3u[0]), float(r3u[1])

            # commission breakdown stats
            cur.execute(
                """SELECT COUNT(DISTINCT platform) AS plat_count,
                          COALESCE(SUM(ABS(gp_amount)), 0) AS total_comm
                   FROM public.rider_deliveries
                   WHERE delivery_date BETWEEN %s AND %s""",
                (first, last),
            )
            r4 = cur.fetchone()
            plat_count, total_comm = int(r4[0]), float(r4[1])

    finally:
        conn.close()

    # Rough ZIP size estimate: ~50KB per Excel file × 3
    zip_est = 150 * 1024

    return {
        "month": month,
        "daybook": {
            "rows": daybook_rows,
            "total_income": round(total_income, 2),
            "total_expense": round(total_expense, 2),
        },
        "pnd3": {
            "rows": pnd3_rows,
            "total_withholding": round(total_wht, 2),
        },
        "category_summary": {
            "categories": cat_count,
            "total_spend": round(total_spend, 2),
            "uncategorized_rows": uncat_count,
            "uncategorized_total": round(uncat_total, 2),
        },
        "commission_breakdown": {
            "platforms": plat_count,
            "total_commission": round(total_comm, 2),
        },
        "zip_bundle": {
            "files": 3,
            "size_bytes_est": zip_est,
        },
    }


def _missing_transfer_slips(
    statement_rows: list[dict],
    daybook_rows: list[dict],
) -> list[dict]:
    daybook_expense_statement_ids = {
        str(row["ref_id"])
        for row in daybook_rows
        if row["direction"] == "expense"
        and row["source"] != "pos_cashflow"
        and row["ref_id"] is not None
    }
    return [
        row
        for row in statement_rows
        if row["id"] in daybook_expense_statement_ids
        and (row["debit"] or 0) > 0
        and not row.get("matched_invoice_is_credit_card", False)
        and not row["has_slip"]
    ]


def _query_readiness_facts(
    first: date,
    last: date,
    branch_code: str,
) -> tuple[dict, dict]:
    """Acquire Phase A readiness facts with six batched SELECTs plus close checks."""
    conn = get_db_conn()
    try:
        conn.set_session(readonly=True, isolation_level="REPEATABLE READ")
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    COUNT(*) AS row_count,
                    COUNT(DISTINCT import_batch_id) AS batch_count,
                    MIN(txn_date) AS first_observed_date,
                    MAX(txn_date) AS last_observed_date,
                    COUNT(*) FILTER (WHERE match_status = 'needs_review') AS review_count,
                    COALESCE(
                        SUM(ABS(amount)) FILTER (WHERE match_status = 'needs_review'),
                        0
                    ) AS review_amount
                FROM public.bank_statement_entries
                WHERE txn_date BETWEEN %s AND %s
                  AND branch_code = %s
                """,
                (first, last, branch_code),
            )
            statement_stats = dict(
                zip((column[0] for column in cur.description), cur.fetchone())
            )

            cur.execute(
                """
                SELECT
                    COUNT(*) FILTER (
                        WHERE direction = 'expense' AND category_code IS NULL
                    ) AS uncategorized_count,
                    COALESCE(
                        SUM(amount) FILTER (
                            WHERE direction = 'expense' AND category_code IS NULL
                        ),
                        0
                    ) AS uncategorized_amount,
                    COALESCE(
                        SUM(amount) FILTER (WHERE direction = 'income'),
                        0
                    ) AS income_total,
                    COALESCE(
                        SUM(amount) FILTER (WHERE direction = 'expense'),
                        0
                    ) AS expense_total,
                    COALESCE(
                        SUM(amount) FILTER (
                            WHERE direction = 'expense' AND category_code IS NOT NULL
                        ),
                        0
                    ) AS categorized_expense
                FROM public.v_daybook_pnl
                WHERE entry_date BETWEEN %s AND %s
                  AND (branch_code = %s OR branch_code IS NULL)
                """,
                (first, last, branch_code),
            )
            pnl_stats = dict(
                zip((column[0] for column in cur.description), cur.fetchone())
            )

            cur.execute(
                """
                SELECT
                    b.id::text, b.import_batch_id::text, b.txn_date,
                    b.description, b.debit, b.credit, b.amount,
                    b.category_code, b.source_type, b.match_status,
                    b.matched_invoice_id::text,
                    COALESCE(
                        matched_vb.payment_type = 'credit_card'
                        OR matched_vb.payment_status = 'credit_card',
                        FALSE
                    ) AS matched_invoice_is_credit_card,
                    EXISTS (
                      SELECT 1
                      FROM public.slips s
                      WHERE s.matched_statement_id = b.id
                    ) AS has_slip
                FROM public.bank_statement_entries b
                LEFT JOIN public.vendor_bills matched_vb
                  ON matched_vb.id = b.matched_invoice_id
                WHERE b.txn_date BETWEEN %s AND %s
                  AND b.branch_code = %s
                ORDER BY b.txn_date, b.id
                """,
                (first, last, branch_code),
            )
            statement_rows = _rows_to_dicts(cur)

            cur.execute(
                """
                SELECT
                    scoped.id, scoped.transfer_date, scoped.amount,
                    scoped.raw_image_url, scoped.matched_statement_id,
                    scoped.matched_invoice_id, scoped.match_status,
                    scoped.is_branch_linked, scoped.is_month_unattributed
                FROM (
                    SELECT
                        s.id::text, s.transfer_date, s.amount, s.raw_image_url,
                        s.matched_statement_id::text,
                        s.matched_invoice_id::text,
                        s.match_status,
                        (
                            EXISTS (
                                SELECT 1
                                FROM public.bank_statement_entries b
                                WHERE b.id = s.matched_statement_id
                                  AND b.txn_date BETWEEN %s AND %s
                                  AND b.branch_code = %s
                            )
                            OR EXISTS (
                                SELECT 1
                                FROM public.vendor_bills vb
                                WHERE vb.id = s.matched_invoice_id
                                  AND vb.bill_date BETWEEN %s AND %s
                                  AND COALESCE(vb.branch_code, %s) = %s
                                  AND COALESCE(vb.review_status, '') <> 'rejected'
                            )
                        ) AS is_branch_linked,
                        (
                            s.matched_statement_id IS NULL
                            AND s.matched_invoice_id IS NULL
                            AND s.transfer_date BETWEEN %s AND %s
                        ) AS is_month_unattributed
                    FROM public.slips s
                ) AS scoped
                WHERE scoped.is_branch_linked OR scoped.is_month_unattributed
                ORDER BY scoped.transfer_date, scoped.id
                """,
                (
                    first, last, branch_code,
                    first, last, "thawi_watthana", branch_code,
                    first, last,
                ),
            )
            slip_rows = _rows_to_dicts(cur)

            cur.execute(
                """
                SELECT
                    vb.id::text, vb.bill_date, vb.amount, vb.category_code,
                    vb.review_status, vb.payment_type, vb.payment_status,
                    vb.attachment_url,
                    a.id::text AS attachment_id, a.page_no, a.file_url
                FROM public.vendor_bills vb
                LEFT JOIN public.attachments a
                  ON a.parent_type = 'vendor_bill' AND a.parent_id = vb.id
                WHERE vb.bill_date BETWEEN %s AND %s
                  AND COALESCE(vb.branch_code, %s) = %s
                  AND COALESCE(vb.review_status, '') <> 'rejected'
                ORDER BY vb.id, a.page_no, a.id
                """,
                (first, last, "thawi_watthana", branch_code),
            )
            bill_rows = _rows_to_dicts(cur)

            cur.execute(
                """
                SELECT
                    entry_date, direction, amount, category_code, source,
                    ref_id::text, label, counterparty
                FROM public.v_daybook_pnl
                WHERE entry_date BETWEEN %s AND %s
                  AND (branch_code = %s OR branch_code IS NULL)
                ORDER BY entry_date, source, ref_id
                """,
                (first, last, branch_code),
            )
            daybook_rows = _rows_to_dicts(cur)

        from monthly_close_routes import run_all_checks

        detected_risks = run_all_checks(conn, first, last, branch_code)
        danger_risks = [
            risk for risk in detected_risks if risk.get("severity") == "danger"
        ]
    finally:
        conn.close()

    bill_groups: dict[str, dict] = {}
    attachment_url_failures = 0
    for row in bill_rows:
        bill_id = row["id"]
        group = bill_groups.setdefault(
            bill_id,
            {
                "amount": row["amount"] or 0,
                "review_status": row["review_status"],
                "payment_type": row["payment_type"],
                "payment_status": row["payment_status"],
                "attachment_url": row["attachment_url"],
                "has_attachment_row": False,
                "has_page_url": False,
            },
        )
        if row["attachment_id"] is not None:
            group["has_attachment_row"] = True
            if row["file_url"]:
                group["has_page_url"] = True
            else:
                attachment_url_failures += 1

    def has_invoice_page(group: dict) -> bool:
        return bool(group["has_page_url"] or group["attachment_url"])

    missing_bills = [
        group for group in bill_groups.values() if not has_invoice_page(group)
    ]
    confirmed_cards = [
        group
        for group in bill_groups.values()
        if group["review_status"] == "confirmed"
        and (
            group["payment_type"] == "credit_card"
            or group["payment_status"] == "credit_card"
        )
    ]
    missing_confirmed_cards = [
        group for group in confirmed_cards if not has_invoice_page(group)
    ]
    missing_transfer_slips = _missing_transfer_slips(statement_rows, daybook_rows)
    branch_linked_slip_rows = [
        row for row in slip_rows if row["is_branch_linked"]
    ]
    month_unattributed_slip_rows = [
        row for row in slip_rows if row["is_month_unattributed"]
    ]

    drift = (
        (pnl_stats["expense_total"] or 0)
        - (pnl_stats["categorized_expense"] or 0)
        - (pnl_stats["uncategorized_amount"] or 0)
    )
    danger_amount = sum((risk.get("amount") or 0 for risk in danger_risks), 0)
    missing_bill_amount = sum((group["amount"] for group in missing_bills), 0)
    missing_card_amount = sum(
        (group["amount"] for group in missing_confirmed_cards), 0
    )
    missing_transfer_amount = sum(
        ((row["debit"] or 0) for row in missing_transfer_slips), 0
    )
    missing_bill_without_attachment_row = sum(
        1
        for group in bill_groups.values()
        if not group["has_attachment_row"] and not group["attachment_url"]
    )
    image_url_failure_count = (
        sum(1 for row in branch_linked_slip_rows if not row["raw_image_url"])
        + attachment_url_failures
        + missing_bill_without_attachment_row
    )

    facts = {
        "statement": {
            "row_count": int(statement_stats["row_count"] or 0),
            "batch_count": int(statement_stats["batch_count"] or 0),
            "first_observed_date": statement_stats["first_observed_date"],
            "last_observed_date": statement_stats["last_observed_date"],
            "declared_period_verified": False,
        },
        "statement_needs_review": {
            "count": int(statement_stats["review_count"] or 0),
            "amount": statement_stats["review_amount"] or 0,
        },
        "uncategorized": {
            "count": int(pnl_stats["uncategorized_count"] or 0),
            "amount": pnl_stats["uncategorized_amount"] or 0,
        },
        "reconciliation": {
            "verified": False,
            "basis": "no_independent_comparator",
            "internal_partition": {
                "ok": abs(float(drift)) <= 0.01,
                "drift": drift,
            },
        },
        "monthly_close_dangers": {
            "count": len(danger_risks),
            "amount": danger_amount,
        },
        "transfer_evidence": {
            "missing_slip_count": len(missing_transfer_slips),
            "missing_slip_amount": missing_transfer_amount,
        },
        "invoice_evidence": {
            "missing_page_count": len(missing_bills),
            "missing_page_amount": missing_bill_amount,
        },
        "credit_cards": {
            "confirmed_count": len(confirmed_cards),
            "missing_invoice_page_count": len(missing_confirmed_cards),
            "missing_invoice_page_amount": missing_card_amount,
        },
        "image_url_failures": {"count": image_url_failure_count},
    }
    fingerprint_inputs = {
        "month": first.strftime("%Y-%m"),
        "branch_code": branch_code,
        "daybook_rows": daybook_rows,
        "statement_rows": statement_rows,
        "branch_linked_slip_rows": branch_linked_slip_rows,
        "month_unattributed_slip_rows": month_unattributed_slip_rows,
        "bills_and_pages": bill_rows,
        "monthly_close": [
            {
                "risk_key": risk["risk_key"],
                "severity": risk["severity"],
                "amount": risk.get("amount"),
            }
            for risk in danger_risks
        ],
    }
    return facts, fingerprint_inputs


@router.get("/readiness")
def export_readiness(
    month: str = Query(..., description="YYYY-MM"),
    branch_code: Literal["thawi_watthana"] = Query("thawi_watthana"),
    _admin: dict = Depends(_require_admin_role),
):
    first, last = _readiness_month_range(month)
    facts, fingerprint_inputs = _query_readiness_facts(first, last, branch_code)
    payload = build_readiness(
        month,
        branch_code,
        {
            **facts,
            "today": bkk_today(),
            "month_start": first,
            "month_end": last,
        },
    )
    payload["generated_at"] = bkk_now().isoformat()
    payload["source_fingerprint"] = {
        "version": "phase-a-v1",
        "sha256": canonical_fingerprint(fingerprint_inputs),
    }
    return payload


_AUDIT_BRANCH_CODE = "thawi_watthana"


def _query_audit_cash_basis_rows(cur, first: date, last: date) -> tuple[list[dict], list[str]]:
    """Return all non-petty cash-basis expenses and only their real Statement IDs."""
    cur.execute(
        """SELECT d.entry_date, d.amount, d.category_code,
                  COALESCE(ec.name_th, d.category_code, 'ไม่ระบุ') AS category_name_th,
                  d.counterparty, d.label, d.source,
                  d.ref_id::text AS ref_id,
                  b.id::text AS statement_id
           FROM public.v_daybook_pnl d
           LEFT JOIN public.expense_categories ec ON ec.code = d.category_code
           LEFT JOIN public.bank_statement_entries b
             ON b.id::text = d.ref_id::text
           WHERE d.direction = 'expense'
             AND d.source <> 'pos_cashflow'
             AND d.entry_date BETWEEN %s AND %s
           ORDER BY d.entry_date, d.source, d.ref_id""",
        (first, last),
    )
    rows = _rows_to_dicts(cur)
    statement_ids = list(dict.fromkeys(
        str(row["statement_id"])
        for row in rows
        if row.get("statement_id") is not None
    ))
    return rows, statement_ids


def _group_invoice_evidence(rows: list[dict], signer) -> dict[str, dict]:
    """Group and sign invoice pages without downloading storage objects."""
    groups: dict[str, dict] = {}
    for row in rows:
        group_key = row.get("stmt_id") or row.get("source_id")
        if group_key is None:
            continue
        key = str(group_key)
        group = groups.setdefault(
            key,
            {
                "invoice_id": str(row.get("invoice_id") or row.get("source_id") or ""),
                "invoice_no": row.get("invoice_no"),
                "vendor_name": row.get("vendor_name"),
                "payment_type": row.get("payment_type"),
                "payment_status": row.get("payment_status"),
                "attachment_url": row.get("attachment_url"),
                "_has_attachment_row": False,
                "_attachments": [],
            },
        )
        for field in (
            "invoice_id", "invoice_no", "vendor_name", "payment_type",
            "payment_status", "attachment_url",
        ):
            if not group.get(field) and row.get(field):
                group[field] = str(row[field]) if field == "invoice_id" else row[field]

        has_attachment_row = (
            row.get("attachment_id") is not None
            or row.get("page_no") is not None
            or row.get("file_url") is not None
        )
        if has_attachment_row:
            group["_has_attachment_row"] = True
        if row.get("file_url"):
            group["_attachments"].append(
                (row.get("page_no"), str(row.get("attachment_id") or ""), row["file_url"])
            )

    for group in groups.values():
        ordered = sorted(
            group.pop("_attachments"),
            key=lambda item: (
                item[0] is None,
                int(item[0]) if item[0] is not None else 0,
                item[1],
            ),
        )
        pages = []
        seen_urls = set()
        next_page = max(
            (int(page_no) for page_no, _, _ in ordered if page_no is not None),
            default=0,
        ) + 1
        for page_no, _, raw_url in ordered:
            if raw_url in seen_urls:
                continue
            seen_urls.add(raw_url)
            normalized_page = int(page_no) if page_no is not None else next_page
            if page_no is None:
                next_page += 1
            pages.append({"page_no": normalized_page, "image_url": signer(raw_url)})
        if (
            not group.pop("_has_attachment_row")
            and group.get("attachment_url")
            and group["attachment_url"] not in seen_urls
        ):
            pages.append({
                "page_no": 1,
                "image_url": signer(group["attachment_url"]),
            })
        group.pop("attachment_url", None)
        group["pages"] = pages
        # Deployed schema-v1 frontend reads this primary-image alias.
        group["image_url"] = pages[0]["image_url"] if pages else None
    return groups


def _payment_method(invoice: dict | None, statement_id: str | None) -> str:
    if invoice and (
        invoice.get("payment_type") == "credit_card"
        or invoice.get("payment_status") == "credit_card"
    ):
        return "Credit Card"
    if statement_id is not None:
        return "Bank Transfer"
    if invoice and invoice.get("payment_type") == "transfer":
        return "Bank Transfer"
    if invoice and invoice.get("payment_type") == "cash":
        return "Cash"
    return "Other"


def _public_invoice_evidence(
    invoice: dict | None,
    *,
    invoice_id: str | None = None,
    invoice_no=None,
    vendor_name=None,
) -> dict | None:
    """Return only fields in the public schema-v2 invoice evidence contract."""
    if invoice is None and invoice_id is None:
        return None
    source = invoice or {}
    pages = [
        {
            "page_no": page.get("page_no"),
            "image_url": page.get("image_url"),
        }
        for page in (source.get("pages") or [])
    ]
    image_url = source.get("image_url")
    if image_url is None and pages:
        image_url = pages[0]["image_url"]
    return {
        "invoice_id": source.get("invoice_id") or invoice_id,
        "invoice_no": source.get("invoice_no") or invoice_no,
        "vendor_name": source.get("vendor_name") or vendor_name,
        "pages": pages,
        # Deployed schema-v1 frontend reads this primary-image alias.
        "image_url": image_url,
    }


def _assemble_audit_vouchers(
    *,
    transfer_rows: list[dict],
    card_rows: list[dict],
    slips_by_stmt: dict,
    inv_by_stmt: dict,
    card_invoices: dict,
    wht_rules: dict,
) -> list[dict]:
    """Assemble cash-basis vouchers first, then supplementary card evidence."""
    vouchers: list[dict] = []
    for row in transfer_rows:
        category_code = row.get("category_code")
        rule = wht_rules.get(category_code) if category_code else None
        amount = float(row.get("amount") or 0)
        wht = None
        if rule:
            wht = {
                "rate": rule["wht_pct"],
                "amount": round(amount * rule["wht_pct"] / 100.0, 2),
            }
        raw_source_id = row.get("ref_id")
        source_id = (
            str(raw_source_id)
            if raw_source_id is not None
            else (
                f"{row.get('source') or 'unknown'}:"
                f"{row.get('entry_date')}:{row.get('amount') or 0}"
            )
        )
        statement_id = (
            str(row["statement_id"])
            if row.get("statement_id") is not None
            else None
        )
        invoice_source = inv_by_stmt.get(statement_id)
        payment_method = _payment_method(invoice_source, statement_id)
        label = row.get("label") or ""
        vouchers.append({
            "seq": len(vouchers) + 1,
            "source_type": "cash_basis_expense",
            "source_id": source_id,
            "statement_id": statement_id,
            "payment_method": payment_method,
            "requires_slip": payment_method == "Bank Transfer",
            "cash_basis_included": True,
            # Schema-v1 compatibility fields consumed by the deployed frontend.
            "date": str(row["entry_date"]),
            "counterparty": row.get("counterparty") or label or "",
            "description": label,
            "category_code": category_code,
            "category_name_th": row.get("category_name_th") or "ไม่ระบุ",
            "amount": round(amount, 2),
            "wht": wht,
            "slip": (
                None
                if payment_method == "Credit Card"
                else slips_by_stmt.get(statement_id)
            ),
            "invoice": _public_invoice_evidence(invoice_source),
        })

    for row in card_rows:
        source_id = str(row["source_id"])
        invoice = _public_invoice_evidence(
            card_invoices.get(source_id),
            invoice_id=source_id,
            invoice_no=row.get("invoice_no"),
            vendor_name=row.get("vendor_name"),
        )
        amount = float(row.get("amount") or 0)
        vouchers.append({
            "seq": len(vouchers) + 1,
            "source_type": "credit_card_invoice",
            "source_id": source_id,
            "statement_id": None,
            "payment_method": "Credit Card",
            "requires_slip": False,
            "cash_basis_included": False,
            # Schema-v1 compatibility fields consumed by the deployed frontend.
            "date": str(row["entry_date"]),
            "counterparty": row.get("vendor_name") or "",
            "description": row.get("label") or row.get("vendor_name") or "",
            "category_code": row.get("category_code"),
            "category_name_th": row.get("category_name_th") or "ไม่ระบุ",
            "amount": round(amount, 2),
            "wht": None,
            "slip": None,
            "invoice": invoice,
        })
    return vouchers


def _expenses_without_required_slip(month: str, vouchers: list[dict]) -> list[dict]:
    return [
        {
            "pv": f"PV-{month.replace('-', '')}-{voucher['seq']:03d}",
            "date": voucher["date"],
            "counterparty": voucher["counterparty"],
            "description": voucher["description"],
            "amount": voucher["amount"],
        }
        for voucher in vouchers
        if voucher["requires_slip"] and voucher["slip"] is None
    ]


def _audit_summary(
    *,
    cash_basis_expense: float,
    petty_total: float,
    vouchers: list[dict],
) -> dict:
    cash_basis_vouchers = [
        voucher for voucher in vouchers if voucher["cash_basis_included"]
    ]
    supplementary_cards = [
        voucher
        for voucher in vouchers
        if not voucher["cash_basis_included"]
        and voucher["payment_method"] == "Credit Card"
    ]
    cash_basis_total = round(
        sum(float(voucher["amount"]) for voucher in cash_basis_vouchers), 2
    )
    supplementary_total = round(
        sum(float(voucher["amount"]) for voucher in supplementary_cards), 2
    )
    expense_pnl = round(float(cash_basis_expense), 2)
    return {
        "expense_pnl": expense_pnl,
        # Deployed schema-v1 frontend treats vouchers as cash-basis bank payments.
        "voucher_count": len(cash_basis_vouchers),
        "evidence_voucher_count": len(vouchers),
        "cash_basis_voucher_count": len(cash_basis_vouchers),
        "cash_basis_voucher_total": cash_basis_total,
        "supplementary_credit_card_count": len(supplementary_cards),
        "supplementary_credit_card_total": supplementary_total,
        "cash_basis_reconciliation_drift": round(
            cash_basis_total + float(petty_total) - expense_pnl, 2
        ),
        # Deployed schema-v1 frontend expects voucher_total to reconcile to P&L.
        "voucher_total": cash_basis_total,
    }


def _build_audit_package_payload(
    *,
    month: str,
    generated_at: str,
    income_pnl: float,
    expense_pnl: float,
    vouchers: list[dict],
    petty_cash: list[dict],
    bills_without_attachment: list[dict],
    unmatched_slips: list[dict],
) -> dict:
    petty_total = round(
        sum(float(row["amount"]) for row in petty_cash), 2
    )
    cash_basis_vouchers = [
        voucher for voucher in vouchers if voucher["cash_basis_included"]
    ]
    expenses_without_slip = _expenses_without_required_slip(
        month, cash_basis_vouchers
    )
    summary = _audit_summary(
        cash_basis_expense=expense_pnl,
        petty_total=petty_total,
        vouchers=vouchers,
    )
    summary.update({
        "income_pnl": round(float(income_pnl), 2),
        "petty_count": len(petty_cash),
        "petty_total": petty_total,
        "missing_counts": {
            "no_slip": len(expenses_without_slip),
            "no_invoice_attachment": len(bills_without_attachment),
            "unmatched_slips": len(unmatched_slips),
        },
    })
    return {
        "schema_version": 2,
        "month": month,
        "month_label_th": _month_label_th(month),
        "generated_at": generated_at,
        "summary": summary,
        # Backend-first compatibility: legacy collection stays cash-basis-only.
        "vouchers": cash_basis_vouchers,
        "evidence_vouchers": vouchers,
        "petty_cash": petty_cash,
        "missing": {
            "expenses_without_slip": expenses_without_slip,
            "bills_without_attachment": bills_without_attachment,
            "unmatched_slips": unmatched_slips,
        },
    }


@router.get("/audit-package")
def export_audit_package(month: str = Query(..., description="YYYY-MM")):
    """ชุดเอกสารตรวจสอบรายเดือน (ใบสำคัญจ่าย + เงินสดย่อย + รายการรอเอกสาร).

    Read-only JSON bundle for the printable A4 audit-package page.
    Voucher numbering is stateless (PV-YYYYMM-### by entry_date,ref_id) — the
    printed/archived PDF is the immutable snapshot (design review 2026-07-13, 5b)."""
    from tax_routes import WHT_RULES  # noqa: PLC0415 — single source of WHT rates
    # The `uploads` storage bucket is private (security hardening 2026-05-31, GAP 2)
    # — stored .../object/public/... paths 404 unless signed at read time. Reuse the
    # existing helper (see slip_routes.py for the same lazy-import pattern) instead
    # of returning raw DB URLs, which would leave every <img> broken.
    try:
        from main import _sign_uploads_url  # noqa: PLC0415
    except Exception:
        def _sign_uploads_url(url, expires_in: int = 86400):  # type: ignore[no-redef]
            return url

    first, last = _month_range(month)
    conn = get_db_conn()
    try:
        conn.set_session(readonly=True)
        with conn.cursor() as cur:
            # ── Summary (same P&L basis as daybook export) ──
            cur.execute(
                """SELECT COALESCE(SUM(CASE WHEN direction='income'  THEN amount END),0),
                          COALESCE(SUM(CASE WHEN direction='expense' THEN amount END),0)
                   FROM public.v_daybook_pnl
                   WHERE entry_date BETWEEN %s AND %s""",
                (first, last),
            )
            _s = cur.fetchone()
            income_pnl, expense_pnl = float(_s[0] or 0), float(_s[1] or 0)

            # ── Voucher rows: bank-side expenses (petty cash goes to its own book) ──
            vrows, statement_ids = _query_audit_cash_basis_rows(cur, first, last)

            # ── Evidence: slips matched to those statement rows ──
            slips_by_stmt: dict = {}
            if statement_ids:
                cur.execute(
                    """SELECT matched_statement_id::text AS stmt_id, raw_image_url, ref_no,
                              transfer_date, transfer_time
                       FROM public.slips
                       WHERE matched_statement_id::text = ANY(%s)
                       ORDER BY matched_statement_id, transfer_date, id""",
                    (statement_ids,),
                )
                for s in _rows_to_dicts(cur):
                    slips_by_stmt[s["stmt_id"]] = {
                        "image_url": _sign_uploads_url(s["raw_image_url"]),
                        "ref_no": s["ref_no"],
                        "transfer_date": str(s["transfer_date"]) if s["transfer_date"] else None,
                        "transfer_time": str(s["transfer_time"]) if s["transfer_time"] else None,
                    }

            # ── Evidence: every invoice page linked to those statement rows ──
            linked_invoice_rows: list[dict] = []
            if statement_ids:
                cur.execute(
                    """SELECT
                           b.id::text AS stmt_id, vb.id::text AS invoice_id,
                           vb.attachment_url, vb.invoice_no, vb.vendor_name,
                           vb.payment_type, vb.payment_status,
                           a.id::text AS attachment_id, a.page_no, a.file_url
                       FROM public.bank_statement_entries b
                       JOIN public.vendor_bills vb ON vb.id = b.matched_invoice_id
                       LEFT JOIN public.attachments a
                         ON a.parent_type = 'vendor_bill' AND a.parent_id = vb.id
                       WHERE b.id::text = ANY(%s)
                       ORDER BY b.id, a.page_no, a.id""",
                    (statement_ids,),
                )
                linked_invoice_rows = _rows_to_dicts(cur)
            inv_by_stmt = _group_invoice_evidence(
                linked_invoice_rows, signer=_sign_uploads_url
            )

            # ── Supplementary evidence: confirmed card bills not already linked
            #    to an in-scope Statement. One batched query includes every page.
            cur.execute(
                """SELECT
                       vb.id::text AS source_id, vb.bill_date AS entry_date,
                       vb.vendor_name, vb.invoice_no, vb.amount, vb.category_code,
                       COALESCE(ec.name_th, vb.category_code, 'ไม่ระบุ') AS category_name_th,
                       vb.attachment_url, a.id::text AS attachment_id,
                       a.page_no, a.file_url
                   FROM public.vendor_bills vb
                   LEFT JOIN public.expense_categories ec ON ec.code=vb.category_code
                   LEFT JOIN public.attachments a
                     ON a.parent_type='vendor_bill' AND a.parent_id=vb.id
                   WHERE vb.review_status='confirmed'
                     AND vb.bill_date BETWEEN %s AND %s
                     AND COALESCE(vb.branch_code, %s) = %s
                     AND (vb.payment_type='credit_card' OR vb.payment_status='credit_card')
                     AND NOT EXISTS (
                         SELECT 1 FROM public.bank_statement_entries b
                         WHERE b.matched_invoice_id=vb.id
                           AND b.txn_date BETWEEN %s AND %s
                           AND b.branch_code=%s
                     )
                   ORDER BY vb.bill_date, vb.id, a.page_no, a.id""",
                (
                    first, last, _AUDIT_BRANCH_CODE, _AUDIT_BRANCH_CODE,
                    first, last, _AUDIT_BRANCH_CODE,
                ),
            )
            card_page_rows = _rows_to_dicts(cur)
            card_invoices = _group_invoice_evidence(
                card_page_rows, signer=_sign_uploads_url
            )
            card_rows: list[dict] = []
            seen_card_ids: set[str] = set()
            for row in card_page_rows:
                source_id = str(row["source_id"])
                if source_id not in seen_card_ids:
                    seen_card_ids.add(source_id)
                    card_rows.append(row)

            vouchers = _assemble_audit_vouchers(
                transfer_rows=vrows,
                card_rows=card_rows,
                slips_by_stmt=slips_by_stmt,
                inv_by_stmt=inv_by_stmt,
                card_invoices=card_invoices,
                wht_rules=WHT_RULES,
            )

            # ── Petty cash book (pos_cashflow) ──
            cur.execute(
                """SELECT d.entry_date, d.label,
                          COALESCE(ec.name_th, d.category_code, 'ไม่ระบุ') AS category_name_th,
                          d.amount
                   FROM public.v_daybook_pnl d
                   LEFT JOIN public.expense_categories ec ON ec.code = d.category_code
                   WHERE d.direction = 'expense'
                     AND d.source = 'pos_cashflow'
                     AND d.entry_date BETWEEN %s AND %s
                   ORDER BY d.entry_date, d.ref_id""",
                (first, last),
            )
            petty = [
                {"date": str(p["entry_date"]), "description": p["label"] or "",
                 "category_name_th": p["category_name_th"], "amount": round(float(p["amount"] or 0), 2)}
                for p in _rows_to_dicts(cur)
            ]
            # ── Missing-documents schedule (3 explicit types — design review risk #2) ──
            cur.execute(
                """SELECT bill_date, vendor_name, invoice_no, amount
                   FROM public.vendor_bills vb
                   WHERE COALESCE(vb.review_status, '') <> 'rejected'
                     AND vb.attachment_url IS NULL
                     AND NOT EXISTS (
                         SELECT 1
                         FROM public.attachments a
                         WHERE a.parent_type='vendor_bill'
                           AND a.parent_id=vb.id
                           AND a.file_url IS NOT NULL
                     )
                     AND vb.bill_date BETWEEN %s AND %s
                     AND COALESCE(vb.branch_code, %s) = %s
                   ORDER BY vb.bill_date, vb.id""",
                (first, last, _AUDIT_BRANCH_CODE, _AUDIT_BRANCH_CODE),
            )
            bills_without_attachment = [
                {"date": str(b["bill_date"]), "vendor_name": b["vendor_name"],
                 "invoice_no": b["invoice_no"], "amount": round(float(b["amount"] or 0), 2)}
                for b in _rows_to_dicts(cur)
            ]
            cur.execute(
                """SELECT transfer_date, amount, memo, recipient_name
                   FROM public.slips
                   WHERE matched_statement_id IS NULL
                     AND transfer_date BETWEEN %s AND %s
                   ORDER BY transfer_date, id""",
                (first, last),
            )
            unmatched_slips = [
                {"date": str(s["transfer_date"]), "amount": round(float(s["amount"] or 0), 2),
                 "memo": s["memo"] or "", "recipient_name": s["recipient_name"] or ""}
                for s in _rows_to_dicts(cur)
            ]
    finally:
        conn.close()

    return _build_audit_package_payload(
        month=month,
        generated_at=bkk_now().isoformat(timespec="seconds"),
        income_pnl=income_pnl,
        expense_pnl=expense_pnl,
        vouchers=vouchers,
        petty_cash=petty,
        bills_without_attachment=bills_without_attachment,
        unmatched_slips=unmatched_slips,
    )


@router.get("/health")
def export_health():
    """Smoke test"""
    return {"status": "ok", "endpoints": [
        "/export/summary?month=YYYY-MM",
        "/export/category-summary?month=YYYY-MM",
        "/export/daybook?month=YYYY-MM",
        "/export/pnd3?month=YYYY-MM",
        "/export/commission-breakdown?month=YYYY-MM",
        "/export/zip-bundle?month=YYYY-MM",
    ]}
