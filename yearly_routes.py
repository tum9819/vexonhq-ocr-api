"""
VEXONHQ Phase 18 — Yearly Report
=================================
Annual P&L summary with a compatibility tombstone for the retired P.N.D. export.

Endpoints:
  GET /pnl/yearly            — monthly breakdown for a year (JSON)
  GET /export/yearly         — download Annual P&L Excel
  GET /export/pnd3-annual    — HTTP 410 compatibility tombstone

In main.py add:
    from yearly_routes import router as yearly_router
    app.include_router(yearly_router)
"""

from __future__ import annotations

import io
import logging
import os
from datetime import date, datetime
from typing import Any, Optional
from uuid import UUID

import psycopg2
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

try:
    from main import get_db_conn  # type: ignore
except ImportError:
    def get_db_conn():
        return psycopg2.connect(os.environ["DATABASE_URL"])

log = logging.getLogger("yearly_routes")
router = APIRouter(tags=["yearly"])

DEFAULT_BRANCH = "thawi_watthana"


# ─────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────

def _rows_to_dicts(cur) -> list[dict]:
    if cur.description is None:
        return []
    cols = [d[0] for d in cur.description]
    rows = []
    for r in cur.fetchall():
        row: dict[str, Any] = {}
        for k, v in zip(cols, r):
            if isinstance(v, UUID):
                row[k] = str(v)
            elif isinstance(v, (datetime, date)):
                row[k] = v.isoformat()
            elif hasattr(v, "__float__") and not isinstance(v, (int, float, bool)):
                row[k] = float(v)
            else:
                row[k] = v
        rows.append(row)
    return rows


TH_MONTHS = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
              "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]


# ─────────────────────────────────────────────────────────
# GET /pnl/yearly   — JSON summary for a year
# ─────────────────────────────────────────────────────────

@router.get("/pnl/yearly")
def pnl_yearly(
    year: int = Query(2026, ge=2020, le=2099),
    branch: str = Query(DEFAULT_BRANCH),
):
    """
    Full-year P&L: monthly breakdown + totals + best/worst month.
    Used by /yearly frontend page.
    """
    commission_map = {}  # Will be populated from rider_deliveries
    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            # ── All rollup numbers from v_daybook_pnl (single source of truth) ──
            # Audit C4 fix (2026-05-27): sales_net + rider_net previously came
            # from raw pos_sales_daily / rider_deliveries while income/expense
            # came from v_daybook — they could silently disagree (sales+rider !=
            # income_total). Pulling all four from the same pre-filtered view
            # guarantees they reconcile by construction. v_daybook_pnl already
            # excludes equity/transfer sources, so no inline NOT IN needed.
            cur.execute(
                """SELECT EXTRACT(MONTH FROM entry_date)::int AS m,
                          COALESCE(SUM(CASE WHEN source='pos_sale'
                                           THEN amount ELSE 0 END), 0)::numeric AS sales_net,
                          COALESCE(SUM(CASE WHEN source IN ('rider_income_grab','rider_income_lineman')
                                           THEN amount ELSE 0 END), 0)::numeric AS rider_net,
                          COALESCE(SUM(CASE WHEN direction='income'
                                           THEN amount ELSE 0 END), 0)::numeric AS income_total,
                          COALESCE(SUM(CASE WHEN direction='expense'
                                           THEN amount ELSE 0 END), 0)::numeric AS expense_total
                   FROM public.v_daybook_pnl
                   WHERE branch_code = %s
                     AND EXTRACT(YEAR FROM entry_date) = %s
                   GROUP BY 1""",
                (branch, year),
            )
            daybook_map = {r[0]: (float(r[1] or 0), float(r[2] or 0),
                                  float(r[3] or 0), float(r[4] or 0))
                           for r in cur.fetchall()}

            # ── POS sales bill_count only (v_daybook_pnl doesn't carry it) ────
            cur.execute(
                """SELECT EXTRACT(MONTH FROM sales_date)::int AS m,
                          SUM(bill_count)::int                AS bill_count
                   FROM public.pos_sales_daily
                   WHERE branch_code = %s
                     AND EXTRACT(YEAR FROM sales_date) = %s
                   GROUP BY 1""",
                (branch, year),
            )
            sales_bill_map = {r[0]: int(r[1] or 0) for r in cur.fetchall()}

            # ── Expense bill count ─────────────────────────────────────────────
            cur.execute(
                """SELECT EXTRACT(MONTH FROM bill_date)::int AS m,
                          COUNT(*)::int AS bill_count
                   FROM public.vendor_bills
                   WHERE review_status = 'confirmed'
                     AND bill_date IS NOT NULL
                     AND COALESCE(branch_code, %s) = %s
                     AND EXTRACT(YEAR FROM bill_date) = %s
                   GROUP BY 1""",
                (branch, branch, year),
            )
            exp_bills_map = {r[0]: int(r[1] or 0) for r in cur.fetchall()}

            # ── Commission breakdown (Grab vs Lineman) ────────────────────────
            cur.execute(
                """SELECT EXTRACT(MONTH FROM delivery_date)::int AS m,
                          platform,
                          COALESCE(SUM(gross_sales), 0)::numeric AS gross,
                          COALESCE(SUM(ABS(gp_amount)), 0)::numeric AS commission,
                          COALESCE(SUM(ABS(promo_store)), 0)::numeric AS promo,
                          COALESCE(SUM(net_payout), 0)::numeric AS net,
                          COALESCE(SUM(order_count), 0)::int AS orders
                   FROM public.rider_deliveries
                   WHERE EXTRACT(YEAR FROM delivery_date) = %s
                   GROUP BY 1, 2
                   ORDER BY 1, 2""",
                (year,),
            )
            commission_rows = cur.fetchall()
            commission_map = {}
            for r in commission_rows:
                m, platform = r[0], r[1]
                if m not in commission_map:
                    commission_map[m] = {}
                platform_key = (platform or "unknown").lower()
                commission_map[m][platform_key] = {
                    "gross": float(r[2] or 0),
                    "commission": float(r[3] or 0),
                    "promo": float(r[4] or 0),
                    "net": float(r[5] or 0),
                    "orders": int(r[6] or 0),
                }

    finally:
        conn.close()

    rows = []
    totals = dict(sales_net=0.0, rider_net=0.0, income_total=0.0,
                  expense_total=0.0, gross_profit=0.0, bill_count=0, expense_bill_count=0)

    for m in range(1, 13):
        s_net, r_net, income, expense = daybook_map.get(m, (0.0, 0.0, 0.0, 0.0))
        s_bills = sales_bill_map.get(m, 0)
        e_bills = exp_bills_map.get(m, 0)
        profit = income - expense
        margin = round(profit / income * 100, 1) if income else None

        row = {
            "month": m,
            "month_label": TH_MONTHS[m],
            "year_month": f"{year}-{m:02d}",
            "sales_net": round(s_net, 2),
            "rider_net": round(r_net, 2),
            "income_total": round(income, 2),
            "expense_total": round(expense, 2),
            "gross_profit": round(profit, 2),
            "gross_margin_pct": margin,
            "sales_bill_count": s_bills,
            "expense_bill_count": e_bills,
            "has_data": income > 0 or expense > 0,
        }
        rows.append(row)
        totals["sales_net"] += s_net
        totals["rider_net"] += r_net
        totals["income_total"] += income
        totals["expense_total"] += expense
        totals["gross_profit"] += profit
        totals["bill_count"] += s_bills
        totals["expense_bill_count"] += e_bills

    totals["gross_margin_pct"] = round(
        totals["gross_profit"] / totals["income_total"] * 100, 1
    ) if totals["income_total"] else None

    # Best / worst month (by profit, only months with data)
    data_rows = [r for r in rows if r["has_data"] and r["income_total"] > 0]
    best_month  = max(data_rows, key=lambda r: r["gross_profit"], default=None)
    worst_month = min(data_rows, key=lambda r: r["gross_profit"], default=None)

    return {
        "year": year,
        "branch": branch,
        "months": rows,
        "totals": totals,
        "best_month": best_month,
        "worst_month": worst_month,
        "data_months": len(data_rows),
        "commission": commission_map,
    }


# ─────────────────────────────────────────────────────────
# GET /export/yearly  — download Annual P&L Excel
# ─────────────────────────────────────────────────────────

@router.get("/export/yearly")
def export_yearly(
    year: int = Query(2026, ge=2020, le=2099),
    branch: str = Query(DEFAULT_BRANCH),
):
    """Download full-year P&L as Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    # Reuse pnl_yearly data
    data = pnl_yearly(year=year, branch=branch)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"P&L {year}"

    LEFT   = Alignment(horizontal="left",   vertical="center")
    CENTER = Alignment(horizontal="center",  vertical="center")
    RIGHT  = Alignment(horizontal="right",   vertical="center")

    FONT_TITLE = Font(name="TH Sarabun New", bold=True, size=16)
    FONT_HDR   = Font(name="TH Sarabun New", bold=True, size=12, color="FFFFFF")
    FONT_BODY  = Font(name="TH Sarabun New", size=12)
    FONT_TOTAL = Font(name="TH Sarabun New", bold=True, size=12)

    FILL_HDR   = PatternFill("solid", fgColor="4F46E5")
    FILL_TOTAL = PatternFill("solid", fgColor="E0E7FF")
    FILL_ALT   = PatternFill("solid", fgColor="F5F5FF")

    NUM_FMT = '#,##0'

    def cell(row, col, val, font=None, align=None, fill=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:   c.font      = font
        if align:  c.alignment = align
        if fill:   c.fill      = fill
        if num_fmt: c.number_format = num_fmt
        return c

    # Title
    ws.merge_cells("A1:H1")
    cell(1, 1, f"รายงาน P&L ประจำปี {year} — ร้านสถานีหม่าล่า", FONT_TITLE, CENTER)
    ws.row_dimensions[1].height = 32

    # Header row
    headers = ["เดือน", "ยอดขาย POS", "ยอด Rider", "รวมรายรับ",
               "ค่าใช้จ่าย", "กำไรขั้นต้น", "มาร์จิน %", "จำนวนบิล"]
    for c_idx, h in enumerate(headers, 1):
        c = cell(3, c_idx, h, FONT_HDR, CENTER, FILL_HDR)
    ws.row_dimensions[3].height = 22

    # Data rows
    for i, row in enumerate(data["months"], 1):
        r = 3 + i
        fill = FILL_ALT if i % 2 == 0 else None
        has = row["has_data"]
        cell(r, 1, row["month_label"], FONT_BODY, CENTER, fill)
        cell(r, 2, row["sales_net"]    if has else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell(r, 3, row["rider_net"]    if has else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell(r, 4, row["income_total"] if has else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell(r, 5, row["expense_total"] if has else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell(r, 6, row["gross_profit"] if has else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell(r, 7, row["gross_margin_pct"] if has else None, FONT_BODY, CENTER, fill,
             '0.0"%"')
        cell(r, 8, row["sales_bill_count"] if has else None, FONT_BODY, CENTER, fill)
        ws.row_dimensions[r].height = 20

    # Total row
    t = data["totals"]
    tr = 3 + 13
    cell(tr, 1, "รวมทั้งปี", FONT_TOTAL, CENTER, FILL_TOTAL)
    cell(tr, 2, t["sales_net"],     FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell(tr, 3, t["rider_net"],     FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell(tr, 4, t["income_total"],  FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell(tr, 5, t["expense_total"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell(tr, 6, t["gross_profit"],  FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell(tr, 7, t["gross_margin_pct"], FONT_TOTAL, CENTER, FILL_TOTAL, '0.0"%"')
    cell(tr, 8, t["bill_count"],    FONT_TOTAL, CENTER, FILL_TOTAL)
    ws.row_dimensions[tr].height = 24

    # Best/worst
    if data["best_month"]:
        bm = data["best_month"]
        cell(tr+2, 1, f"📈 เดือนที่กำไรสูงสุด: {bm['month_label']} (฿{bm['gross_profit']:,.0f})",
             FONT_BODY, LEFT)
        ws.merge_cells(f"A{tr+2}:D{tr+2}")
    if data["worst_month"]:
        wm = data["worst_month"]
        cell(tr+3, 1, f"📉 เดือนที่กำไรต่ำสุด: {wm['month_label']} (฿{wm['gross_profit']:,.0f})",
             FONT_BODY, LEFT)
        ws.merge_cells(f"A{tr+3}:D{tr+3}")

    # Column widths
    for c_idx, w in enumerate([12, 16, 14, 16, 16, 16, 12, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 2: Commission Breakdown (Grab vs Lineman)
    # ──────────────────────────────────────────────────────────────────────
    ws_comm = wb.create_sheet("Commission Breakdown")

    def cell_comm(row, col, val, font=None, align=None, fill=None, num_fmt=None):
        c = ws_comm.cell(row=row, column=col, value=val)
        if font:   c.font      = font
        if align:  c.alignment = align
        if fill:   c.fill      = fill
        if num_fmt: c.number_format = num_fmt
        return c

    # Title
    ws_comm.merge_cells("A1:I1")
    cell_comm(1, 1, f"รายงานการหักค่าคอมมิชชัน ประจำปี {year}", FONT_TITLE, CENTER)
    ws_comm.row_dimensions[1].height = 32

    # Header row
    comm_headers = ["เดือน", "ยอดขาย Gross",
                    "ค่าคอม Grab", "ส่วนลด Grab",
                    "ค่าคอม Lineman", "ส่วนลด Lineman",
                    "รวมค่าคอม", "รวมส่วนลด", "ยอดขาย Net"]
    for c_idx, h in enumerate(comm_headers, 1):
        cell_comm(3, c_idx, h, FONT_HDR, CENTER, FILL_HDR)
    ws_comm.row_dimensions[3].height = 22

    # Commission data rows
    comm_map = data.get("commission", {})
    comm_totals = {
        "gross": 0.0, "comm_grab": 0.0, "promo_grab": 0.0,
        "comm_lineman": 0.0, "promo_lineman": 0.0, "net": 0.0
    }

    for i, m in enumerate(range(1, 13), 1):
        r = 3 + i
        month_data = comm_map.get(m, {})
        fill = FILL_ALT if i % 2 == 0 else None

        grab_data = month_data.get("grab", {})
        lineman_data = month_data.get("lineman", {})

        gross_m = (grab_data.get("gross", 0) or 0) + (lineman_data.get("gross", 0) or 0)
        comm_grab = grab_data.get("commission", 0) or 0
        promo_grab = grab_data.get("promo", 0) or 0
        comm_lineman = lineman_data.get("commission", 0) or 0
        promo_lineman = lineman_data.get("promo", 0) or 0
        net_m = (grab_data.get("net", 0) or 0) + (lineman_data.get("net", 0) or 0)
        total_comm = comm_grab + comm_lineman
        total_promo = promo_grab + promo_lineman

        cell_comm(r, 1, TH_MONTHS[m], FONT_BODY, CENTER, fill)
        cell_comm(r, 2, gross_m if gross_m > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell_comm(r, 3, comm_grab if comm_grab > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell_comm(r, 4, promo_grab if promo_grab > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell_comm(r, 5, comm_lineman if comm_lineman > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell_comm(r, 6, promo_lineman if promo_lineman > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell_comm(r, 7, total_comm if total_comm > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell_comm(r, 8, total_promo if total_promo > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        cell_comm(r, 9, net_m if net_m > 0 else None, FONT_BODY, RIGHT, fill, NUM_FMT)
        ws_comm.row_dimensions[r].height = 20

        comm_totals["gross"] += gross_m
        comm_totals["comm_grab"] += comm_grab
        comm_totals["promo_grab"] += promo_grab
        comm_totals["comm_lineman"] += comm_lineman
        comm_totals["promo_lineman"] += promo_lineman
        comm_totals["net"] += net_m

    # Total row
    tr_comm = 3 + 13
    cell_comm(tr_comm, 1, "รวมทั้งปี", FONT_TOTAL, CENTER, FILL_TOTAL)
    cell_comm(tr_comm, 2, comm_totals["gross"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell_comm(tr_comm, 3, comm_totals["comm_grab"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell_comm(tr_comm, 4, comm_totals["promo_grab"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell_comm(tr_comm, 5, comm_totals["comm_lineman"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell_comm(tr_comm, 6, comm_totals["promo_lineman"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell_comm(tr_comm, 7, comm_totals["comm_grab"] + comm_totals["comm_lineman"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell_comm(tr_comm, 8, comm_totals["promo_grab"] + comm_totals["promo_lineman"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    cell_comm(tr_comm, 9, comm_totals["net"], FONT_TOTAL, RIGHT, FILL_TOTAL, NUM_FMT)
    ws_comm.row_dimensions[tr_comm].height = 24

    # Column widths
    for c_idx, w in enumerate([12, 16, 14, 14, 14, 14, 14, 14, 14], 1):
        ws_comm.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"annual_pnl_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────
# GET /export/pnd3-annual — retired compatibility endpoint
# ─────────────────────────────────────────────────

@router.get("/export/pnd3-annual")
def export_pnd3_annual(
    year: Optional[str] = Query(None, description="retired"),
):
    """Retired: no annual P.N.D. workbook is generated from expense categories."""
    del year
    from tax_routes import raise_wht_decommissioned
    raise_wht_decommissioned()
