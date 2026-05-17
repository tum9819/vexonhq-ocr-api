"""
tax_routes.py — Phase 34: ภ.ง.ด.3/53 Withholding Tax Monthly Report
=======================================================================
Endpoints:
  GET  /tax/wht-summary?month=YYYY-MM          — WHT data for the month
  GET  /tax/wht-export?month=YYYY-MM           — Download XLSX for accountant

WHT Categories (from bank_statement_entries):
  musician_fee  → มาตรา 40(8)  → 3% WHT
  rent          → มาตรา 40(5)  → 5% WHT

Usage:
  รายจ่ายที่นายจ้างต้องหักภาษี ณ ที่จ่าย และนำส่งสรรพากรภายในวันที่ 7 ของเดือนถัดไป
"""

import io
import logging
from datetime import date
from calendar import monthrange
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

import psycopg2
import psycopg2.extras

try:
    from main import get_db_conn  # type: ignore
except ImportError:
    def get_db_conn():  # type: ignore
        raise RuntimeError("get_db_conn not available")

logger = logging.getLogger("tax")

router = APIRouter()

# ── WHT rate config ──────────────────────────────────────────────────────────

WHT_RULES = {
    "musician_fee": {
        "label":       "ค่าดนตรี / นักดนตรี",
        "section":     "มาตรา 40(8)",
        "wht_pct":     3.0,
        "form":        "ภ.ง.ด.3",
    },
    "rent": {
        "label":       "ค่าเช่า",
        "section":     "มาตรา 40(5)",
        "wht_pct":     5.0,
        "form":        "ภ.ง.ด.3",
    },
    "service_fee": {
        "label":       "ค่าบริการ / ค่าจ้าง",
        "section":     "มาตรา 40(6)",
        "wht_pct":     3.0,
        "form":        "ภ.ง.ด.3",
    },
}

MONTH_NAMES_TH = [
    "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
    "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม",
]


def _month_bounds(month: Optional[str]):
    """Return (start_date, end_date) for a YYYY-MM string."""
    if not month:
        today = date.today()
        month = f"{today.year}-{today.month:02d}"
    try:
        y, m = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        raise HTTPException(400, "month must be YYYY-MM")
    last_day = monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, last_day)


def _format_month_th(month: str) -> str:
    y, m = int(month[:4]), int(month[5:7])
    return f"{MONTH_NAMES_TH[m]} {y + 543}"


# ── GET /tax/wht-summary ──────────────────────────────────────────────────────

@router.get("/tax/wht-summary")
def wht_summary(
    month: Optional[str] = Query(None, description="YYYY-MM (default: current month)"),
    branch_code: str = Query("thawi_watthana"),
):
    """
    WHT summary for the selected month.
    Returns category totals + individual transaction rows.
    """
    start, end = _month_bounds(month)
    if not month:
        month = start.strftime("%Y-%m")

    categories = list(WHT_RULES.keys())

    conn = get_db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    category_code,
                    txn_date,
                    description,
                    debit::numeric AS amount
                FROM public.bank_statement_entries
                WHERE branch_code = %s
                  AND txn_date BETWEEN %s AND %s
                  AND category_code = ANY(%s)
                  AND debit > 0
                ORDER BY category_code, txn_date
                """,
                (branch_code, start, end, categories),
            )
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    finally:
        conn.close()

    # Build category summaries
    summaries: dict[str, dict] = {}
    transactions: list[dict] = []

    for row in rows:
        cat = row["category_code"]
        rule = WHT_RULES[cat]
        amount = float(row["amount"] or 0)
        wht = round(amount * rule["wht_pct"] / 100, 2)
        net_paid = round(amount - wht, 2)

        # Per-transaction entry
        transactions.append({
            "category_code":  cat,
            "label":          rule["label"],
            "section":        rule["section"],
            "form":           rule["form"],
            "wht_pct":        rule["wht_pct"],
            "txn_date":       str(row["txn_date"]),
            "description":    row["description"],
            "amount_paid":    amount,   # amount actually debited from bank
            "wht_amount":     wht,
            "net_before_wht": round(amount / (1 - rule["wht_pct"] / 100), 2),  # gross estimate
        })

        # Accumulate category totals
        if cat not in summaries:
            summaries[cat] = {
                "category_code": cat,
                "label":         rule["label"],
                "section":       rule["section"],
                "form":          rule["form"],
                "wht_pct":       rule["wht_pct"],
                "txn_count":     0,
                "total_paid":    0.0,
                "total_wht":     0.0,
            }
        summaries[cat]["txn_count"] += 1
        summaries[cat]["total_paid"] = round(summaries[cat]["total_paid"] + amount, 2)
        summaries[cat]["total_wht"] = round(summaries[cat]["total_wht"] + wht, 2)

    summary_list = sorted(summaries.values(), key=lambda x: x["category_code"])
    total_wht = round(sum(s["total_wht"] for s in summary_list), 2)
    total_paid = round(sum(s["total_paid"] for s in summary_list), 2)

    # Due date: 7th of next month
    y, m = int(month[:4]), int(month[5:7])
    nm = m + 1 if m < 12 else 1
    ny = y if m < 12 else y + 1
    due_date = f"{ny}-{nm:02d}-07"
    due_date_th = f"7 {MONTH_NAMES_TH[nm]} {ny + 543}"

    return {
        "month":         month,
        "month_th":      _format_month_th(month),
        "branch_code":   branch_code,
        "due_date":      due_date,
        "due_date_th":   due_date_th,
        "total_paid":    total_paid,
        "total_wht":     total_wht,
        "total_net":     round(total_paid - total_wht, 2),
        "summary":       summary_list,
        "transactions":  transactions,
        "note": (
            "ยอดภาษีที่คำนวณจากยอดที่ตัดบัญชีจริง "
            "กรุณาตรวจสอบกับบัญชีและนำส่งสรรพากรภายใน " + due_date_th
        ),
    }


# ── GET /tax/wht-export ───────────────────────────────────────────────────────

@router.get("/tax/wht-export")
def wht_export(
    month: Optional[str] = Query(None, description="YYYY-MM"),
    branch_code: str = Query("thawi_watthana"),
):
    """Export WHT report as .xlsx for accountant filing."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    data = wht_summary(month=month, branch_code=branch_code)
    month_label = data["month_th"]
    month_str = data["month"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ภงด3_{month_str}"

    # ── Styles ──
    GREEN  = "1A6B3C"
    LGRAY  = "F5F5F5"
    HEADER = PatternFill("solid", fgColor=GREEN)
    ALT    = PatternFill("solid", fgColor=LGRAY)
    bold   = Font(name="TH Sarabun New", bold=True, size=13)
    header_font = Font(name="TH Sarabun New", bold=True, size=12, color="FFFFFF")
    normal = Font(name="TH Sarabun New", size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right", vertical="center")
    left   = Alignment(horizontal="left", vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"รายงานภาษีหัก ณ ที่จ่าย (ภ.ง.ด.3) — {month_label}"
    ws["A1"].font = Font(name="TH Sarabun New", bold=True, size=16)
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = f"ร้านมาราสเตชั่น  |  ครบกำหนดนำส่ง: {data['due_date_th']}"
    ws["A2"].font = Font(name="TH Sarabun New", size=12, color="555555")
    ws["A2"].alignment = center

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # ── Summary section ──
    ws["A4"] = "สรุปยอดภาษีหัก ณ ที่จ่าย"
    ws["A4"].font = bold
    ws["A4"].alignment = left

    sum_headers = ["หมวด", "มาตรา", "แบบ", "อัตรา", "จำนวนรายการ", "ยอดที่จ่าย (บาท)", "ภาษีที่ต้องนำส่ง (บาท)"]
    sum_cols    = ["A", "B", "C", "D", "E", "F", "G"]

    row = 5
    for col, hdr in zip(sum_cols, sum_headers):
        cell = ws[f"{col}{row}"]
        cell.value = hdr
        cell.font = header_font
        cell.fill = HEADER
        cell.alignment = center
        cell.border = border

    row = 6
    for i, s in enumerate(data["summary"]):
        fill = ALT if i % 2 else PatternFill()
        for col, val in zip(sum_cols, [
            s["label"], s["section"], s["form"],
            f'{s["wht_pct"]:.0f}%',
            s["txn_count"],
            s["total_paid"],
            s["total_wht"],
        ]):
            c = ws[f"{col}{row}"]
            c.value = val
            c.font = normal
            c.alignment = right if col in ("E","F","G","D") else left
            c.border = border
            c.fill = fill
            if col in ("F", "G"):
                c.number_format = '#,##0.00'
        row += 1

    # Total row
    for col, val in zip(sum_cols, ["รวมทั้งสิ้น", "", "", "", "", data["total_paid"], data["total_wht"]]):
        c = ws[f"{col}{row}"]
        c.value = val
        c.font = Font(name="TH Sarabun New", bold=True, size=12)
        c.alignment = right if col in ("E","F","G") else left
        c.fill = PatternFill("solid", fgColor="E8F5E9")
        c.border = border
        if col in ("F", "G"):
            c.number_format = '#,##0.00'

    row += 2

    # ── Transaction detail section ──
    ws[f"A{row}"] = "รายละเอียดรายการ"
    ws[f"A{row}"].font = bold
    ws[f"A{row}"].alignment = left
    row += 1

    txn_headers = ["วันที่", "รายละเอียด", "หมวด", "มาตรา", "อัตรา", "ยอดที่จ่าย (บาท)", "ภาษีหัก ณ ที่จ่าย (บาท)"]
    for col, hdr in zip(sum_cols, txn_headers):
        c = ws[f"{col}{row}"]
        c.value = hdr
        c.font = header_font
        c.fill = HEADER
        c.alignment = center
        c.border = border
    row += 1

    for i, t in enumerate(data["transactions"]):
        fill = ALT if i % 2 else PatternFill()
        for col, val in zip(sum_cols, [
            t["txn_date"],
            t["description"],
            t["label"],
            t["section"],
            f'{t["wht_pct"]:.0f}%',
            t["amount_paid"],
            t["wht_amount"],
        ]):
            c = ws[f"{col}{row}"]
            c.value = val
            c.font = normal
            c.alignment = right if col in ("E","F","G") else left
            c.border = border
            c.fill = fill
            if col in ("F", "G"):
                c.number_format = '#,##0.00'
        row += 1

    # Grand total row
    for col, val in zip(sum_cols, ["", "รวมภาษีที่ต้องนำส่ง", "", "", "", data["total_paid"], data["total_wht"]]):
        c = ws[f"{col}{row}"]
        c.value = val
        c.font = Font(name="TH Sarabun New", bold=True, size=12)
        c.alignment = right if col in ("F","G") else left
        c.fill = PatternFill("solid", fgColor="FFFDE7")
        c.border = border
        if col in ("F", "G"):
            c.number_format = '#,##0.00'

    # ── Column widths ──
    col_widths = [14, 42, 20, 16, 10, 20, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Note ──
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = f"หมายเหตุ: {data['note']}"
    ws[f"A{row}"].font = Font(name="TH Sarabun New", size=11, color="888888", italic=True)
    ws[f"A{row}"].alignment = left

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"wht_report_{month_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
